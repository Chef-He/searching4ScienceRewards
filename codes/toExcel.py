import openpyxl
import os
from openpyxl.styles import Alignment
from typing import List, Dict

def toexcel(datas: List[Dict], file: str):
    if not datas:
        return True
    items = 0
    file_exists = os.path.exists(file)

    if file_exists:
        wb = openpyxl.load_workbook(file, read_only=False)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["省份", "年份", "项目", "获奖人员", "主要完成单位", "提名单位", "奖励类型", "奖励级别"])  
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            cell = ws[f"{col}1"]
            cell.alignment = Alignment(horizontal='center')

    try:
        for data in datas:
            ws.append([
                data.get("prov", ""),  
                data.get("year", ""),      
                data.get("proj", ""),    
                data.get("name", ""),    
                data.get("unit",""),
                data.get("non_unit", ""),
                data.get("type", ""),
                data.get("level", "")
            ])
            items += 1

        column_widths = {
        'D': 50,
        'C': 50, 
        'E': 25
    }
    
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width

        wb.save(file)
        print(f"{items}条数据已写入至 {file}")
        return False   
        
    except Exception as e:
        print(f"写入数据时发生错误: {str(e)}")
        return True
    finally:
        wb.close()
