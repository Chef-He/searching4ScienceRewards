import openpyxl
import os
from openpyxl.styles import Alignment
from typing import List, Dict

def toexcel(datas: List[Dict], file: str):
    """将数据写入 Excel 文件"""
    # 检查文件是否存在
    items = 0
    file_exists = os.path.exists(file)

    # 创建一个工作簿
    if file_exists:
        # 如果文件存在，加载工作簿
        wb = openpyxl.load_workbook(file)
        ws = wb.active
    else:
        # 如果文件不存在，创建一个新的工作簿和表头
        wb = openpyxl.Workbook()
        ws = wb.active
        # 写入表头
        ws.append(["省份", "年份", "项目", "获奖人员", "奖励类型", "奖励级别"])  
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws[f"{col}1"]
            cell.alignment = Alignment(horizontal='center')

    # 写入数据
    try:
        for data in datas:
            ws.append([
                data.get("province", ""),  
                data.get("year", ""),      
                data.get("project", ""),    
                data.get("name_unit", ""),    
                data.get("award_type", ""),
                data.get("award_level", "")
            ])
            items += 1

        # 保存文件
        column_widths = {
        'D': 75,
        'C': 50, 
        'E': 25
    }
    
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width

        wb.save(file)
        print(f"{items}条数据已写入至 {file}")
    except Exception as e:
        print(f"写入数据时发生错误: {str(e)}")
    finally:
        wb.close()